import openpyxl
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui
import os
from random import randint

# Carregando a primeira planilha
workbook = openpyxl.load_workbook('Contatos_Brasil.xlsx')
pagina_clientes = workbook['Sheet1']

url_video = 'https://youtube.com/shorts/s_fi3xGlHAI?feature=share'

linhas_para_excluir = []

mensagem = f'{url_video}\n\n *Olá tudo bem ?*\n🚀 *Venha fazer parte dessa comunidade no Discord !!* 🚀 \n\n🔗 *Link de Convite:* https://discord.com/invite/QXzdGW8FaT \n\n*Aqui você vai encontrar:* \n\n*Vagas de emprego* 💼 , *Freelancer 🛠e Cursos gratuitos* 📚 *disponíveis 24 horas. Além de espaço para desenvolver projetos colaborativos* 🤝 *e fazer networking* 🌐'

mensagem_exterior = f'{url_video}\n\n *Hello, how are you?*\n🚀 *Come and be part of this community on Discord!!* 🚀 \n\n🔗 *Invitation Link:* https://discord.com/invite/QXzdGW8FaT \n\n*Here you will find:* \n\n*Job vacancies* 💼, *Freelancer 🛠and Free courses* 📚 *available 24 hours. In addition to space to develop collaborative projects* 🤝 *and network* 🌐'

# Loop para a primeira planilha
for i, linha in enumerate(pagina_clientes.iter_rows(min_row=2), start=2):
    telefone = linha[0].value
    link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={
        telefone}&text={quote(mensagem)}'
    webbrowser.open(link_mensagem_whatsapp)
    sleep(randint(60, 300))  # Espera de 1 a 5 minutos
    seta = pyautogui.locateCenterOnScreen(
        'seta.png', grayscale=True, confidence=0.8)
    if seta is not None:
        pyautogui.click(seta[0], seta[1])
        sleep(randint(60, 300))  # Espera de 1 a 5 minutos
        pyautogui.hotkey('ctrl', 'w')

        linhas_para_excluir.append(i)
    else:
        print(f'Botão de envio não encontrado para {telefone}')
    sleep(randint(60, 300))  # Espera de 1 a 5 minutos

# Salve a primeira planilha
workbook.save('Contatos_Brasil.xlsx')

# Carregando a segunda planilha
workbook_exterior = openpyxl.load_workbook('Contatos_Exterior.xlsx')
pagina_clientes_exterior = workbook_exterior['Sheet1']

# Loop para a segunda planilha
for i, linha in enumerate(pagina_clientes_exterior.iter_rows(min_row=2), start=2):
    telefone = linha[0].value
    link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={
        telefone}&text={quote(mensagem_exterior)}'
    webbrowser.open(link_mensagem_whatsapp)
    sleep(randint(60, 300))  # Espera de 1 a 5 minutos
    seta = pyautogui.locateCenterOnScreen(
        'seta.png', grayscale=True, confidence=0.8)
    if seta is not None:
        pyautogui.click(seta[0], seta[1])
        sleep(randint(60, 300))  # Espera de 1 a 5 minutos
        pyautogui.hotkey('ctrl', 'w')

        linhas_para_excluir.append(i)
    else:
        print(f'Botão de envio não encontrado para {telefone}')
    sleep(randint(60, 300))  # Espera de 1 a 5 minutos

# Salve a segunda planilha
workbook_exterior.save('Contatos_Exterior.xlsx')
